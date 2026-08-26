#!/usr/bin/env python3
"""Build Watchdog's governed supplement from an official DCA Affordable Housing Dashboard export.

The DCA public page states that the dashboard, unlike the tabular AHMS workbook, includes
all-municipality LMI cost burden and HUD-subsidized-unit data. This parser accepts an
official dashboard export (CSV or XLSX) and emits only those two exact municipality facts.

No broader household cost-burden field is substituted for LMI cost burden. No HUD unit
count is inferred from AHMS project rows. Ambiguous municipality identity fails closed.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
CROSSWALK = ROOT / "property/data/budget-pressure.json"
SOURCE_PAGE = "https://www.nj.gov/dca/dlps/hss/MuniStatusReporting.shtml"
DASHBOARD_SOURCE = "NJ DCA Affordable Housing Reporting Dashboard"

KEYWORDS = {
    "municipality": ["municipality", "municipal name", "town"],
    "county": ["county"],
    "hud_subsidized_units": ["hud subsidized units", "hud-subsidized units", "hud subsidized", "hud units"],
    "low_income_cost_burden": ["lmi cost burden", "low and moderate income cost burden", "low- and moderate-income cost burden", "low income cost burden"],
}
REQUIRED = ["municipality", "hud_subsidized_units", "low_income_cost_burden"]


def clean(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def norm(v: object) -> str:
    s = clean(v).upper().rstrip(".")
    s = re.sub(r"\bTOWNSHIP\b", "TWP", s)
    s = re.sub(r"\bBOROUGH\b", "BORO", s)
    return s


def norm_county(v: object) -> str:
    return re.sub(r"\s+COUNTY$", "", norm(v))


def aliases(v: object) -> list[str]:
    full = norm(v)
    if not full:
        return []
    out = [full]
    for suffix in (" TWP", " BORO", " CITY", " TOWN", " VILLAGE"):
        if full.endswith(suffix):
            out.append(full[: -len(suffix)])
    return list(dict.fromkeys(out))


def number(v: object) -> float | None:
    if v is None or clean(v) in ("", "-", "--", "N/A"):
        return None
    text = clean(v).replace(",", "").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def load_crosswalk():
    data = json.loads(CROSSWALK.read_text(encoding="utf-8"))
    municipalities = data.get("municipalities") or {}
    by_pair: dict[tuple[str, str], str] = {}
    candidates: dict[str, set[str]] = defaultdict(set)
    for district, row in municipalities.items():
        county = norm_county(row.get("county"))
        for alias in aliases(row.get("name")):
            candidates[alias].add(str(district))
            if county:
                by_pair.setdefault((alias, county), str(district))
    unique = {name: next(iter(ids)) for name, ids in candidates.items() if len(ids) == 1}
    return municipalities, by_pair, unique


def header_map(headers: list[object]) -> dict[str, int]:
    cells = [clean(v).lower() for v in headers]
    result: dict[str, int] = {}
    for field, variants in KEYWORDS.items():
        for idx, cell in enumerate(cells):
            if cell and any(variant in cell for variant in variants):
                result[field] = idx
                break
    return result


def rows_from_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    for i, row in enumerate(rows[:30]):
        cols = header_map(row)
        if all(field in cols for field in REQUIRED):
            return row, rows[i + 1 :], cols, f"CSV row {i + 1}"
    raise RuntimeError("Could not find required dashboard-export columns in CSV.")


def rows_from_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        preview = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True))
        for i, row in enumerate(preview, start=1):
            cols = header_map(list(row))
            if all(field in cols for field in REQUIRED):
                body = list(ws.iter_rows(min_row=i + 1, values_only=True))
                return list(row), [list(x) for x in body], cols, f"{ws.title} row {i}"
    raise RuntimeError("Could not find required dashboard-export columns in workbook.")


def resolve(muni: object, county: object, by_pair, unique) -> str | None:
    c = norm_county(county)
    if c:
        for alias in aliases(muni):
            hit = by_pair.get((alias, c))
            if hit:
                return hit
    for alias in aliases(muni):
        hit = unique.get(alias)
        if hit:
            return hit
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("export", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    municipalities, by_pair, unique = load_crosswalk()
    if args.export.suffix.lower() == ".csv":
        _, rows, cols, source_table = rows_from_csv(args.export)
    elif args.export.suffix.lower() in (".xlsx", ".xlsm"):
        _, rows, cols, source_table = rows_from_xlsx(args.export)
    else:
        raise RuntimeError("Dashboard supplement must be CSV or XLSX export.")

    out: dict[str, dict] = {}
    unmatched: list[str] = []
    conflicts: list[str] = []
    for row in rows:
        if cols["municipality"] >= len(row):
            continue
        muni = row[cols["municipality"]]
        if not clean(muni):
            continue
        county = row[cols["county"]] if "county" in cols and cols["county"] < len(row) else None
        district = resolve(muni, county, by_pair, unique)
        if not district:
            unmatched.append(clean(muni))
            continue
        hud = number(row[cols["hud_subsidized_units"]]) if cols["hud_subsidized_units"] < len(row) else None
        burden = number(row[cols["low_income_cost_burden"]]) if cols["low_income_cost_burden"] < len(row) else None
        current = out.get(district)
        candidate = {
            "district": district,
            "name": municipalities.get(district, {}).get("name") or clean(muni),
            "county": municipalities.get(district, {}).get("county") or clean(county),
            "hud_subsidized_units": None if hud is None else round(hud),
            "low_income_cost_burden": None if burden is None else round(burden, 2),
        }
        if current and (current.get("hud_subsidized_units") != candidate.get("hud_subsidized_units") or current.get("low_income_cost_burden") != candidate.get("low_income_cost_burden")):
            conflicts.append(district)
            out.pop(district, None)
            continue
        out[district] = candidate

    if conflicts:
        raise RuntimeError(f"Conflicting duplicate municipality dashboard rows: {sorted(set(conflicts))[:20]}")

    payload = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": DASHBOARD_SOURCE,
        "source_page": SOURCE_PAGE,
        "reporting_period": "February 2026",
        "source_table": source_table,
        "municipalities_matched": len(out),
        "unmatched_count": len(unmatched),
        "unmatched_examples": unmatched[:50],
        "field_contract": {
            "hud_subsidized_units": "Exact municipality HUD-subsidized-unit value exported from the official DCA Affordable Housing Dashboard.",
            "low_income_cost_burden": "Exact municipality LMI cost-burden value exported from the official DCA Affordable Housing Dashboard; not the broader all-household cost-burden metric.",
        },
        "municipalities": out,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    print(f"wrote {len(out)} municipalities to {args.output}; unmatched={len(unmatched)}")


if __name__ == "__main__":
    main()
