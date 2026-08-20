#!/usr/bin/env python3
"""Inspect the official NJ DCA Neighborhood Trends workbook without changing marker states."""
from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

SOURCE = "https://www.nj.gov/dca/dhcr/offices/docs/nrtc/Neighborhood_Trends-Database_2026.xlsx"


def clean(value):
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    return text[:240] if text else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    response = requests.get(SOURCE, timeout=120, headers={"User-Agent": "Watchdog source-contract audit"})
    response.raise_for_status()
    temp = args.output.with_suffix(".source.xlsx")
    temp.parent.mkdir(parents=True, exist_ok=True)
    temp.write_bytes(response.content)

    wb = load_workbook(temp, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        sample = []
        for row in ws.iter_rows(min_row=1, max_row=25, max_col=min(ws.max_column or 1, 60), values_only=True):
            sample.append([clean(v) for v in row])
        sheets.append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sample_first_25_rows": sample,
        })
    wb.close()
    temp.unlink(missing_ok=True)

    payload = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": SOURCE,
        "source_bytes": len(response.content),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }
    args.output.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {args.output} with {len(sheets)} sheet diagnostics")


if __name__ == "__main__":
    main()
