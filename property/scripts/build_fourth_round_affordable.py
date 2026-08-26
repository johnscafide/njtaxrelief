#!/usr/bin/env python3
"""Build a governed NJ DCA Fourth Round municipal affordable-housing snapshot.

The extractor reads only the official DCA calculation workbook Final Summary sheet,
identifies columns from the workbook's own headers, and requires exactly 564 unique
municipalities before an output is publishable. Failed gates still emit diagnostics
so source-shape changes can be investigated without weakening the contract.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

SOURCE_URL = "https://www.nj.gov/dca/dlps/pdf/FourthRoundCalculation_Workbook.xlsx"
SOURCE_PAGE = "https://www.nj.gov/dca/dlps/4th_Round_Numbers.shtml"
EXPECTED_MUNICIPALITIES = 564


def clean(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def number(v: object):
    if v is None or clean(v) == "":
        return None
    try:
        x = float(v)
        return int(x) if x.is_integer() else round(x, 6)
    except (TypeError, ValueError):
        return None


def combined_headers(ws) -> dict[int, str]:
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    width = max(len(r) for r in rows)
    out = {}
    for i in range(width):
        parts = []
        for row in rows:
            if i < len(row):
                text = clean(row[i])
                if text and text not in parts:
                    parts.append(text)
        out[i] = " | ".join(parts)
    return out


def find_col(headers: dict[int, str], patterns: list[str], required=True):
    for idx, text in headers.items():
        low = text.lower()
        if all(re.search(p, low, re.I) for p in patterns):
            return idx
    if required:
        raise RuntimeError(f"Missing required Final Summary header matching {patterns}")
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", type=Path, required=True)
    args = ap.parse_args()

    r = requests.get(SOURCE_URL, timeout=180)
    r.raise_for_status()
    tmp = args.output.with_suffix(".xlsx")
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_bytes(r.content)
    wb = load_workbook(tmp, read_only=True, data_only=True)
    ws = wb["Final Summary"]
    headers = combined_headers(ws)

    c_fips = find_col(headers, [r"county subdivision fips"])
    c_muni = find_col(headers, [r"municipality"])
    c_county = find_col(headers, [r"county"])
    c_region = find_col(headers, [r"region"], required=False)
    c_present = find_col(headers, [r"present need"])
    c_prospective = find_col(headers, [r"prospective need"])

    optional_specs = {
        "land_capacity_factor": [r"land capacity", r"factor"],
        "nonresidential_value_factor": [r"nonresidential", r"factor"],
        "income_capacity_factor": [r"income capacity", r"factor"],
        "qualified_urban_aid": [r"qualified urban aid municipality"],
    }
    optional_cols = {k: find_col(headers, pats, required=False) for k, pats in optional_specs.items()}

    municipalities = {}
    duplicates = []
    skipped_examples = []
    candidate_rows = 0
    for excel_row, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        fips = clean(row[c_fips] if c_fips < len(row) else "")
        muni = clean(row[c_muni] if c_muni < len(row) else "")
        county = clean(row[c_county] if c_county < len(row) else "")
        if not fips or not muni or not county:
            continue
        candidate_rows += 1
        digits = re.sub(r"\D", "", fips)
        if len(digits) not in (5, 10):
            if len(skipped_examples) < 30:
                skipped_examples.append({"row": excel_row, "fips": fips, "municipality": muni, "county": county})
            continue
        # Normalize either 5-digit county-subdivision code or 10-digit state+county+subdivision FIPS.
        key = digits[-5:] if len(digits) == 10 else digits
        if key in municipalities:
            duplicates.append({"fips": key, "row": excel_row, "municipality": muni})
            continue
        rec = {
            "county_subdivision_fips": key,
            "municipality": muni,
            "county": county,
            "region": clean(row[c_region]) if c_region is not None and c_region < len(row) else None,
            "present_need": number(row[c_present] if c_present < len(row) else None),
            "prospective_need": number(row[c_prospective] if c_prospective < len(row) else None),
        }
        for field, col in optional_cols.items():
            if col is not None and col < len(row):
                value = row[col]
                rec[field] = clean(value) if field == "qualified_urban_aid" else number(value)
        municipalities[key] = rec

    missing_present = sum(1 for x in municipalities.values() if x["present_need"] is None)
    missing_prospective = sum(1 for x in municipalities.values() if x["prospective_need"] is None)
    errors = []
    if duplicates:
        errors.append(f"duplicate_fips={len(duplicates)}")
    if len(municipalities) != EXPECTED_MUNICIPALITIES:
        errors.append(f"municipality_count={len(municipalities)} expected={EXPECTED_MUNICIPALITIES}")
    if missing_present or missing_prospective:
        errors.append(f"missing_required present={missing_present} prospective={missing_prospective}")

    payload = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "NJ DCA Fourth Round (2025-2035) Affordable Housing Calculations",
        "source_url": SOURCE_URL,
        "source_page": SOURCE_PAGE,
        "legal_context": "DCA describes these as non-binding calculations/guidance under P.L. 2024, c.2; Watchdog must not present them as legal determinations.",
        "validation": {
            "publishable": not errors,
            "errors": errors,
            "candidate_rows": candidate_rows,
            "municipality_count": len(municipalities),
            "duplicate_count": len(duplicates),
            "duplicate_examples": duplicates[:30],
            "skipped_examples": skipped_examples,
            "required_field_coverage": {
                "present_need": len(municipalities)-missing_present,
                "prospective_need": len(municipalities)-missing_prospective,
            },
        },
        "resolved_columns": {
            "fips": c_fips + 1,
            "municipality": c_muni + 1,
            "county": c_county + 1,
            "region": None if c_region is None else c_region + 1,
            "present_need": c_present + 1,
            "prospective_need": c_prospective + 1,
            **{k: None if v is None else v + 1 for k, v in optional_cols.items()},
        },
        "final_summary_headers": headers,
        "municipalities": municipalities if not errors else {},
    }
    args.output.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    tmp.unlink(missing_ok=True)
    print(json.dumps(payload["validation"], sort_keys=True))
    if errors:
        sys.exit(2)


if __name__ == "__main__":
    main()
