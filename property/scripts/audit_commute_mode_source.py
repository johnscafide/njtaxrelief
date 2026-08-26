#!/usr/bin/env python3
from __future__ import annotations
import argparse,json,re,tempfile
from datetime import datetime,timezone
from pathlib import Path
import requests
from openpyxl import load_workbook
SOURCE_URL='https://www.nj.gov/dca/dhcr/offices/docs/nrtc/Neighborhood_Trends-Database_2026.xlsx'
TERMS=['commut','drove alone','carpool','public transportation','public transit','walking to work','bicycle','worked from home','taxicab','motorcycle','other means']
def clean(v): return re.sub(r'\s+',' ',str(v or '')).strip()
def main():
 p=argparse.ArgumentParser();p.add_argument('--output',type=Path,required=True);a=p.parse_args()
 r=requests.get(SOURCE_URL,timeout=180);r.raise_for_status()
 with tempfile.NamedTemporaryFile(suffix='.xlsx',delete=False) as f:f.write(r.content);tmp=Path(f.name)
 wb=load_workbook(tmp,read_only=True,data_only=True);hits=[]
 for ws in wb.worksheets:
  limit=ws.max_row if ws.title=='Data Dictionary' else (4 if ws.title=='Data by Municipality' else 0)
  if not limit: continue
  for ri,row in enumerate(ws.iter_rows(min_row=1,max_row=limit,values_only=True),1):
   for ci,v in enumerate(row,1):
    t=clean(v);lo=t.lower()
    if t and any(x in lo for x in TERMS):hits.append({'sheet':ws.title,'row':ri,'column':ci,'text':t[:1000]})
 payload={'schema_version':1,'generated_at':datetime.now(timezone.utc).isoformat(),'source_url':SOURCE_URL,'hits':hits,'hit_count':len(hits)}
 a.output.parent.mkdir(parents=True,exist_ok=True);a.output.write_text(json.dumps(payload,indent=2),encoding='utf-8');wb.close();tmp.unlink(missing_ok=True);print(f'commute hits={len(hits)}')
if __name__=='__main__':main()
