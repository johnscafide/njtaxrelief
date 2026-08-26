#!/usr/bin/env python3
from __future__ import annotations
import argparse, json, re, tempfile
from datetime import datetime, timezone
from pathlib import Path
import requests
from openpyxl import load_workbook

SOURCE_URL='https://www.nj.gov/dca/dhcr/offices/docs/nrtc/Neighborhood_Trends-Database_2026.xlsx'
TERMS=['walkability','walking to work','walk to work','median block size','population density']

def clean(v): return re.sub(r'\s+',' ',str(v or '')).strip()

def main():
    p=argparse.ArgumentParser(); p.add_argument('--output',type=Path,required=True); a=p.parse_args()
    r=requests.get(SOURCE_URL,timeout=180); r.raise_for_status()
    with tempfile.NamedTemporaryFile(suffix='.xlsx',delete=False) as f:
        f.write(r.content); tmp=Path(f.name)
    wb=load_workbook(tmp,read_only=True,data_only=True)
    hits=[]
    for ws in wb.worksheets:
        limit=ws.max_row if ws.title in ('Data Dictionary','Introduction','Cover') else min(ws.max_row,12)
        for ri,row in enumerate(ws.iter_rows(min_row=1,max_row=limit,values_only=True),1):
            for ci,value in enumerate(row,1):
                text=clean(value); low=text.lower()
                if text and any(t in low for t in TERMS):
                    hits.append({'sheet':ws.title,'row':ri,'column':ci,'text':text[:1000]})
    ws=wb['Data by Municipality']
    header_window=[]
    for ri,row in enumerate(ws.iter_rows(min_row=1,max_row=4,min_col=82,max_col=92,values_only=True),1):
        header_window.append({'row':ri,'cells':[{'column':ci+82,'text':clean(value)} for ci,value in enumerate(row)]})
    first_municipalities=[]
    for ri,row in enumerate(ws.iter_rows(min_row=4,max_row=8,min_col=1,max_col=92,values_only=True),4):
        first_municipalities.append({'row':ri,'municipality':clean(row[0]),'county':clean(row[1]),'columns_82_92':[clean(v) for v in row[81:92]]})
    payload={'schema_version':2,'generated_at':datetime.now(timezone.utc).isoformat(),'source_url':SOURCE_URL,'source_bytes':len(r.content),'hits':hits,'hit_count':len(hits),'municipality_header_window_82_92':header_window,'municipality_sample_rows':first_municipalities}
    a.output.parent.mkdir(parents=True,exist_ok=True); a.output.write_text(json.dumps(payload,indent=2),encoding='utf-8')
    wb.close(); tmp.unlink(missing_ok=True); print(f"walkability hits={len(hits)}")
if __name__=='__main__': main()
