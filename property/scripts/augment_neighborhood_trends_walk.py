#!/usr/bin/env python3
"""Add the exact DCA 2020-24 Walking-to-Work percentage to Neighborhood Trends.

Fails closed unless the official workbook keeps the exact municipality header pair,
all 564 existing municipality records reconcile uniquely, and known canaries match.
This is a mobility statistic, not a generalized walkability score.
"""
from __future__ import annotations
import argparse,json,re,tempfile
from datetime import datetime,timezone
from pathlib import Path
import requests
from openpyxl import load_workbook

SOURCE='https://www.nj.gov/dca/dhcr/offices/docs/nrtc/Neighborhood_Trends-Database_2026.xlsx'
METRIC='% Walking to Work'; PERIOD='2020-24 Estimate'; EXPECTED=564

def clean(v): return re.sub(r'\s+',' ',str(v or '')).strip()
def norm(v):
    s=clean(v).upper().replace('.','')
    s=re.sub(r'\bTOWNSHIP\b','TWP',s);s=re.sub(r'\bBOROUGH\b','BORO',s)
    return re.sub(r'\s+',' ',s).strip()
def aliases(v):
    s=norm(v);out=[s]
    for suffix in (' TWP',' BORO',' CITY',' TOWN',' VILLAGE'):
        if s.endswith(suffix):out.append(s[:-len(suffix)])
    return list(dict.fromkeys(out))
def county(v): return re.sub(r'\s+COUNTY$','',norm(v))
def pct(v):
    if v is None:return None
    try:x=float(v)
    except (TypeError,ValueError):return None
    return round(x*100 if -1<=x<=1 else x,4)

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--input',type=Path,required=True);ap.add_argument('--output',type=Path,required=True);a=ap.parse_args()
    root=json.loads(a.input.read_text(encoding='utf-8'));existing=root.get('municipalities') or {}
    if len(existing)!=EXPECTED:raise RuntimeError(f'existing municipality gate failed: {len(existing)}')
    pair={};unique={}
    name_candidates={}
    for district,row in existing.items():
        c=county(row.get('county'))
        for name in aliases(row.get('name')):
            if c:pair[(name,c)]=str(district)
            name_candidates.setdefault(name,set()).add(str(district))
    unique={k:next(iter(v)) for k,v in name_candidates.items() if len(v)==1}
    r=requests.get(SOURCE,timeout=180,headers={'User-Agent':'Watchdog governed source builder'});r.raise_for_status()
    with tempfile.NamedTemporaryFile(suffix='.xlsx',delete=False) as f:f.write(r.content);tmp=Path(f.name)
    wb=load_workbook(tmp,read_only=True,data_only=True);ws=wb['Data by Municipality'];it=ws.iter_rows(values_only=True)
    next(it);metrics=list(next(it));periods=list(next(it));current=None;columns={}
    for idx,(m,p) in enumerate(zip(metrics,periods)):
        if clean(m):current=clean(m)
        if current and clean(p):columns[(current,clean(p))]=idx
    if (METRIC,PERIOD) not in columns:raise RuntimeError(f'exact mobility header pair missing: {(METRIC,PERIOD)}')
    col=columns[(METRIC,PERIOD)];seen=set();unmatched=[];missing=[]
    for row in it:
        muni=clean(row[0] if len(row)>0 else '');co=clean(row[1] if len(row)>1 else '')
        if not muni:continue
        district=None
        for n in aliases(muni):
            district=pair.get((n,county(co)))
            if district:break
        if not district:
            for n in aliases(muni):
                district=unique.get(n)
                if district:break
        if not district:
            unmatched.append(f'{muni} ({co})');continue
        if district in seen:raise RuntimeError(f'duplicate municipality match: {district}')
        seen.add(district);value=pct(row[col] if col<len(row) else None)
        if value is None:missing.append(district)
        existing[district]['walking_to_work_share']=value
    wb.close();tmp.unlink(missing_ok=True)
    if len(seen)!=EXPECTED or unmatched:raise RuntimeError(f'municipality reconciliation failed matched={len(seen)} unmatched={unmatched[:10]}')
    # Source canaries from the audited workbook header window.
    if existing['0101'].get('walking_to_work_share')!=1.4275:raise RuntimeError(f"Absecon canary failed: {existing['0101'].get('walking_to_work_share')}")
    if existing['0102'].get('walking_to_work_share')!=11.8692:raise RuntimeError(f"Atlantic City canary failed: {existing['0102'].get('walking_to_work_share')}")
    root['schema_version']=3;root['generated_at']=datetime.now(timezone.utc).isoformat();root['source_bytes']=len(r.content)
    root.setdefault('field_contract',{})['walking_to_work_share']='published 2020-24 % of employed workers age 16 and older walking to work, percentage points; mobility statistic, not a generalized walkability score'
    root.setdefault('excluded_catalog_fields',{})['walkability_score']='Workbook publishes % Walking to Work; Watchdog does not relabel this mobility statistic as a generalized walkability score.'
    root['walking_to_work_validation']={'publishable':True,'municipalities_matched':len(seen),'missing_values':len(missing),'missing_districts':missing,'header_metric':METRIC,'header_period':PERIOD,'canaries':{'0101':1.4275,'0102':11.8692}}
    a.output.write_text(json.dumps(root,separators=(',',':')),encoding='utf-8')
    print(json.dumps(root['walking_to_work_validation'],sort_keys=True))
if __name__=='__main__':main()
