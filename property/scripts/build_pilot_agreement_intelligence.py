#!/usr/bin/env python3
"""Build governed municipal PILOT agreement intelligence from the official NJ DCA workbook.

Input: NJ DCA "PILOT Database and Viewer 2026.xlsx" (2025 UFB submissions).
Output: municipal agreement evidence for Watchdog Data Center.

This parser intentionally distinguishes reported workbook rows from legal conclusions:
- agreement_count is a count of distinct *reported agreement fingerprints*, not a legal count
  of enforceable financial agreements.
- expiration_year / term_remaining are based on the earliest future *reported* agreement
  end date with a usable date, and do not imply that an agreement is active, enforceable,
  amended, or uncancelled.
- project_type is a deterministic type mix across reported rows; no single type is chosen
  when a municipality has multiple reported types.

Missing or ambiguous source values fail closed. Scenario-only PFAF outputs are not parsed.
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
CROSSWALK = ROOT / "property/data/budget-pressure.json"
SOURCE_URL = "https://www.nj.gov/dca/dlgs/misc_docs/2026/PILOT%20Database%20and%20Viewer%202026.xlsx"
SOURCE_LABEL = "NJ DCA PILOT Database and Viewer 2026"
RAW_SHEET_HINT = "raw data from ufbs"

KEYWORDS = {
    "municipality": ["municipality", "municipal name", "town"],
    "county": ["county"],
    "municipal_code": ["municipality code", "municipal code", "muni code", "dlgs code"],
    "project_name": ["project name"],
    "agreement_start_date": ["agreement start date", "start date"],
    "agreement_end_date": ["agreement end date", "end date", "expiration date"],
    "project_type": ["type of project", "project type"],
}
REQUIRED = ["municipality", "project_name"]


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm(value: object) -> str:
    s = clean(value).upper().rstrip(".")
    s = re.sub(r"\bTOWNSHIP\b", "TWP", s)
    s = re.sub(r"\bBOROUGH\b", "BORO", s)
    return s


def norm_county(value: object) -> str:
    return re.sub(r"\s+COUNTY$", "", norm(value))


def aliases(value: object) -> list[str]:
    full = norm(value)
    if not full:
        return []
    out = [full]
    for suffix in (" TWP", " BORO", " CITY", " TOWN", " VILLAGE"):
        if full.endswith(suffix):
            out.append(full[: -len(suffix)])
    return list(dict.fromkeys(out))


def municipal_code(value: object) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    if not digits or len(digits) > 4:
        return None
    return digits.zfill(4)


def iso_date(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    if not text:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def header_map(values: list[object]) -> dict[str, int]:
    cells = [clean(v).lower() for v in values]
    result: dict[str, int] = {}
    for field, variants in KEYWORDS.items():
        for idx, cell in enumerate(cells):
            if cell and any(v in cell for v in variants):
                result[field] = idx
                break
    return result


def load_crosswalk():
    root = json.loads(CROSSWALK.read_text(encoding="utf-8"))
    municipalities = root.get("municipalities") or {}
    by_pair: dict[tuple[str, str], str] = {}
    candidates: dict[str, set[str]] = defaultdict(set)
    for district, row in municipalities.items():
        county = norm_county(row.get("county"))
        for alias in aliases(row.get("name")):
            candidates[alias].add(str(district))
            if county:
                by_pair.setdefault((alias, county), str(district))
    unique = {name: next(iter(ids)) for name, ids in candidates.items() if len(ids) == 1}
    return municipalities, by_pair, unique


def resolve_district(row: list[object], cols: dict[str, int], by_pair, unique) -> tuple[str | None, str]:
    if "municipal_code" in cols:
        code = municipal_code(row[cols["municipal_code"]])
        if code:
            return code, "source_municipal_code"
    muni = row[cols["municipality"]]
    county = row[cols["county"]] if "county" in cols else None
    c = norm_county(county)
    if c:
        for alias in aliases(muni):
            hit = by_pair.get((alias, c))
            if hit:
                return hit, "municipality_county"
    for alias in aliases(muni):
        hit = unique.get(alias)
        if hit:
            return hit, "unique_municipality_name"
    return None, "unmatched"


def fingerprint(project_name: str, start: str | None, end: str | None, project_type: str) -> str:
    return "|".join([norm(project_name), start or "", end or "", norm(project_type)])


def discover_raw_sheet(workbook):
    preferred = [name for name in workbook.sheetnames if RAW_SHEET_HINT in name.lower()]
    ordered = preferred + [name for name in workbook.sheetnames if name not in preferred]
    for name in ordered:
        ws = workbook[name]
        for row_idx in range(1, min(ws.max_row, 50) + 1):
            values = [ws.cell(row_idx, c).value for c in range(1, min(ws.max_column, 80) + 1)]
            cols = header_map(values)
            if all(key in cols for key in REQUIRED) and (
                "agreement_end_date" in cols or "agreement_start_date" in cols or "project_type" in cols
            ):
                return ws, row_idx, cols
    raise RuntimeError("Could not locate the PILOT raw UFB table or required headers; source contract changed.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    municipalities, by_pair, unique = load_crosswalk()
    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    ws, header_row, cols = discover_raw_sheet(wb)

    records: dict[str, dict] = {}
    fingerprints: dict[str, set[str]] = defaultdict(set)
    type_counts: dict[str, Counter[str]] = defaultdict(Counter)
    end_dates: dict[str, set[str]] = defaultdict(set)
    unmatched: list[str] = []
    invalid_end_dates = 0
    rows_seen = 0

    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(values)
        project_name = clean(row[cols["project_name"]]) if cols["project_name"] < len(row) else ""
        muni_name = clean(row[cols["municipality"]]) if cols["municipality"] < len(row) else ""
        if not project_name or not muni_name:
            continue
        rows_seen += 1
        district, method = resolve_district(row, cols, by_pair, unique)
        if not district or district not in municipalities:
            unmatched.append(f"{muni_name}: {project_name}")
            continue

        start = iso_date(row[cols["agreement_start_date"]]) if "agreement_start_date" in cols and cols["agreement_start_date"] < len(row) else None
        raw_end = row[cols["agreement_end_date"]] if "agreement_end_date" in cols and cols["agreement_end_date"] < len(row) else None
        end = iso_date(raw_end)
        if raw_end not in (None, "") and end is None:
            invalid_end_dates += 1
        project_type = clean(row[cols["project_type"]]) if "project_type" in cols and cols["project_type"] < len(row) else ""

        base = municipalities.get(district, {})
        rec = records.setdefault(district, {
            "district": district,
            "name": base.get("name") or muni_name,
            "county": base.get("county"),
            "identity_methods": [],
            "reported_rows": 0,
        })
        if method not in rec["identity_methods"]:
            rec["identity_methods"].append(method)
        rec["reported_rows"] += 1
        fingerprints[district].add(fingerprint(project_name, start, end, project_type))
        if project_type:
            type_counts[district][project_type] += 1
        if end:
            end_dates[district].add(end)

    today = datetime.now(timezone.utc).date()
    for district, rec in records.items():
        rec["reported_agreement_fingerprint_count"] = len(fingerprints[district])
        rec["project_type_mix"] = dict(sorted(type_counts[district].items()))
        dates = sorted(end_dates[district])
        rec["reported_agreement_end_dates"] = dates
        future = [d for d in dates if date.fromisoformat(d) >= today]
        rec["next_reported_expiration_date"] = future[0] if future else None
        rec["next_reported_expiration_year"] = int(future[0][:4]) if future else None

    payload = {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": SOURCE_LABEL,
        "source_url": SOURCE_URL,
        "source_year": 2025,
        "release_year": 2026,
        "source_sheet": ws.title,
        "header_row": header_row,
        "columns_matched": sorted(cols.keys()),
        "rows_seen": rows_seen,
        "municipalities_matched": len(records),
        "unmatched_rows": len(unmatched),
        "unmatched_examples": unmatched[:50],
        "invalid_end_dates": invalid_end_dates,
        "contracts": {
            "agreement_count": "Count of distinct reported fingerprints: normalized project name + reported agreement start date + reported agreement end date + reported project type. Not a legal count of enforceable agreements.",
            "expiration_year": "Earliest future year among valid reported agreement end dates. Not a determination that the agreement remains active or unamended.",
            "term_remaining": "Calculated at request time from the earliest future reported agreement end date; unavailable when no future valid end date exists.",
            "project_type": "Deterministic reported project-type mix; no arbitrary single type is selected when multiple types exist.",
        },
        "municipalities": records,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    wb.close()
    print(f"wrote {len(records)} municipalities from {rows_seen} raw rows to {args.output}")


if __name__ == "__main__":
    main()
