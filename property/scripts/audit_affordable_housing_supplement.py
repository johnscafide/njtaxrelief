#!/usr/bin/env python3
from __future__ import annotations
import argparse, json, re, tempfile
from datetime import datetime, timezone
from pathlib import Path
import requests
from openpyxl import load_workbook

SOURCE_URL = 'https://www.nj.gov/dca/dlps/hss/annualreporting/Affordable%20Housing%20Municipal%20Status%20Report%207-1-26.xlsx'
TARGETS = {
  'affordable_units_pipeline': ['pipeline','under construction','future units','proposed units','units pipeline'],
  'hud_subsidized_units': ['hud subsidized','hud-subsidized','hud assisted','hud-assisted'],
  'low_income_cost_burden': ['lmi cost burden','low income cost burden','low-income cost burden','cost-burdened','cost burden'],
}

def clean(v): return re.sub(r'\s+',' ',str(v or '')).strip()

def main():
    p=argparse.ArgumentParser(); p.add_argument('--output',type=Path,required=True); a=p.parse_args()
    with tempfile.NamedTemporaryFile(suffix='.xlsx',delete=False) as f: tmp=Path(f.name)
    r=requests.get(SOURCE_URL,timeout=120); r.raise_for_status(); tmp.write_bytes(r.content)
    wb=load_workbook(tmp,read_only=True,data_only=True)
    hits={k:[] for k in TARGETS}; sheets=[]
    for ws in wb.worksheets:
        sheets.append({'name':ws.title,'max_row':ws.max_row,'max_column':ws.max_column})
        for row_idx, row in enumerate(ws.iter_rows(min_row=1,max_row=min(ws.max_row,80),values_only=True),1):
            for col_idx, value in enumerate(row,1):
                text=clean(value); low=text.lower()
                if not low: continue
                for key, terms in TARGETS.items():
                    if any(term in low for term in terms):
                        hits[key].append({'sheet':ws.title,'row':row_idx,'column':col_idx,'text':text[:500]})
    payload={'schema_version':1,'generated_at':datetime.now(timezone.utc).isoformat(),'source_url':SOURCE_URL,'workbook_bytes':len(r.content),'sheets':sheets,'target_hits':hits,'target_hit_counts':{k:len(v) for k,v in hits.items()}}
    a.output.parent.mkdir(parents=True,exist_ok=True); a.output.write_text(json.dumps(payload,indent=2),encoding='utf-8')
    wb.close(); tmp.unlink(missing_ok=True)
    print(json.dumps(payload['target_hit_counts']))
if __name__=='__main__': main()
