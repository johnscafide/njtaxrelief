#!/usr/bin/env python3
"""Audit the official NJ DCA Fourth Round affordable-housing calculation workbook.

This source-contract diagnostic downloads the official workbook, records sheet
shape and relevant headers, and traces uncached Final Summary formulas where
openpyxl data_only values are blank. It never publishes Data Center values.
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

SOURCE_URL = "https://www.nj.gov/dca/dlps/pdf/FourthRoundCalculation_Workbook.xlsx"
TARGETS = {
    "municipality": [r"municipalit", r"municipal name"],
    "county": [r"county"],
    "region": [r"region"],
    "present_need": [r"present need"],
    "prospective_need": [r"prospective need"],
    "fair_share": [r"fair share", r"obligation"],
    "land_capacity": [r"land capacity", r"developable land", r"vacant land"],
    "income_capacity": [r"income capacity", r"income factor"],
    "nonresidential": [r"non.?residential", r"nonresidential"],
}


def norm(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    response = requests.get(SOURCE_URL, timeout=120)
    response.raise_for_status()
    content = response.content

    tmp = args.output.with_suffix(".xlsx")
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_bytes(content)
    wb_values = load_workbook(tmp, read_only=True, data_only=True)
    wb_formulas = load_workbook(tmp, read_only=True, data_only=False)

    sheets = []
    hits = {k: [] for k in TARGETS}
    for ws in wb_values.worksheets:
        sheets.append({"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column})
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                text = norm(value)
                if not text:
                    continue
                low = text.lower()
                for key, patterns in TARGETS.items():
                    if any(re.search(p, low, flags=re.I) for p in patterns):
                        hits[key].append({"sheet": ws.title, "row": r_idx, "column": c_idx, "text": text[:500]})

    # Current Final Summary leaf-header contract: C municipality, D county,
    # F present need, G qualified urban aid, L prospective need, Q capped need.
    vws = wb_values["Final Summary"]
    fws = wb_formulas["Final Summary"]
    formula_gap_samples = []
    formula_gap_counts = {"present_need": 0, "prospective_need": 0, "prospective_need_capped": 0}
    for r in range(4, vws.max_row + 1):
        muni = norm(vws.cell(r, 3).value)
        county = norm(vws.cell(r, 4).value)
        if not muni or not county:
            continue
        for key, col in (("present_need", 6), ("prospective_need", 12), ("prospective_need_capped", 17)):
            cached = vws.cell(r, col).value
            if cached is None:
                formula_gap_counts[key] += 1
                if len(formula_gap_samples) < 80:
                    formula_gap_samples.append({
                        "row": r,
                        "municipality": muni,
                        "county": county,
                        "qualified_urban_aid_cached": vws.cell(r, 7).value,
                        "field": key,
                        "cached_value": None,
                        "formula": fws.cell(r, col).value,
                        "present_formula": fws.cell(r, 6).value,
                        "prospective_formula": fws.cell(r, 12).value,
                        "capped_formula": fws.cell(r, 17).value,
                    })

    payload = {
        "schema_version": 2,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_url": SOURCE_URL,
        "workbook_bytes": len(content),
        "sheets": sheets,
        "target_hits": hits,
        "target_hit_counts": {k: len(v) for k, v in hits.items()},
        "formula_gap_counts": formula_gap_counts,
        "formula_gap_samples": formula_gap_samples,
        "activation_gate": {
            "status": "diagnostic_only",
            "rule": "Do not publish Fourth Round markers until municipality-level rows and exact field semantics are validated against DCA methodology and all 564 municipalities are reconciled; blank cached formula cells are not treated as zero."
        }
    }
    args.output.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    tmp.unlink(missing_ok=True)
    print(json.dumps({"targets": payload["target_hit_counts"], "formula_gaps": formula_gap_counts}, sort_keys=True))


if __name__ == "__main__":
    main()
