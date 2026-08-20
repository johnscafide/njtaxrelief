#!/usr/bin/env python3
"""Build governed municipal Neighborhood Trends values from the official 2026 NJ DCA workbook.

Only fields with an exact municipality-level source contract are emitted. The catalog's
commute-mode mix, median real-estate tax, and walkability-score markers are intentionally
not synthesized from unlike fields. A small Municipal Housing Profile compatibility set
is also emitted where the same official DCA workbook publishes the exact municipality
facts or the value is a transparent ratio of those published municipality counts.
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

SOURCE = "https://www.nj.gov/dca/dhcr/offices/docs/nrtc/Neighborhood_Trends-Database_2026.xlsx"
ROOT = Path(__file__).resolve().parents[2]
CROSSWALK = ROOT / "property/data/budget-pressure.json"
CURRENT_LABEL = "2020-24 Estimate"


def norm(value: object) -> str:
    s = re.sub(r"\s+", " ", str(value or "")).strip().upper().rstrip(".")
    s = re.sub(r"\bTOWNSHIP\b", "TWP", s)
    s = re.sub(r"\bBOROUGH\b", "BORO", s)
    s = re.sub(r"\bCITY\b", "CITY", s)
    s = re.sub(r"\bTWP\.?$", "TWP", s)
    s = re.sub(r"\bBORO\.?$", "BORO", s)
    return s


def county_norm(value: object) -> str:
    return re.sub(r"\s+COUNTY$", "", norm(value))


def aliases(value: object) -> list[str]:
    s = norm(value)
    if not s:
        return []
    out = [s]
    for suffix in (" TWP", " BORO", " CITY", " TOWN", " VILLAGE"):
        if s.endswith(suffix):
            out.append(s[: -len(suffix)])
    return list(dict.fromkeys(out))


def num(value):
    if value is None or value == "--" or value == "-":
        return None
    try:
        x = float(value)
    except (TypeError, ValueError):
        return None
    return x


def pct_change(old, new):
    a, b = num(old), num(new)
    if a is None or b is None or a <= 0:
        return None
    return round((b / a - 1) * 100, 2)


def pct_share(numerator, denominator):
    a, b = num(numerator), num(denominator)
    if a is None or b is None or b <= 0:
        return None
    return round(a / b * 100, 2)


def percent_value(value):
    x = num(value)
    if x is None:
        return None
    # Excel percentage cells are commonly stored as fractions while some source
    # tables publish already-scaled percentages. Normalize both to percentage points.
    return round(x * 100 if -1 <= x <= 1 else x, 2)


def rounded(value, digits=2):
    x = num(value)
    return None if x is None else round(x, digits)


def load_crosswalk():
    data = json.loads(CROSSWALK.read_text(encoding="utf-8"))
    rows = data.get("municipalities") or {}
    by_pair: dict[tuple[str, str], str] = {}
    unique_candidates: dict[str, set[str]] = {}
    for district, row in rows.items():
        county = county_norm(row.get("county"))
        for a in aliases(row.get("name")):
            if county:
                by_pair.setdefault((a, county), str(district))
            unique_candidates.setdefault(a, set()).add(str(district))
    unique = {a: next(iter(ids)) for a, ids in unique_candidates.items() if len(ids) == 1}
    return rows, by_pair, unique


def clean_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    response = requests.get(SOURCE, timeout=120, headers={"User-Agent": "Watchdog governed source builder"})
    response.raise_for_status()
    temp = args.output.with_suffix(".source.xlsx")
    temp.parent.mkdir(parents=True, exist_ok=True)
    temp.write_bytes(response.content)

    wb = load_workbook(temp, read_only=True, data_only=True)
    ws = wb["Data by Municipality"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # section labels
    metric_row = list(next(rows_iter))
    period_row = list(next(rows_iter))

    metric = None
    columns: dict[tuple[str, str], int] = {}
    for idx, (m, p) in enumerate(zip(metric_row, period_row)):
        if clean_header(m):
            metric = clean_header(m)
        period = clean_header(p)
        if metric and period:
            columns[(metric, period)] = idx

    required = {
        "land_area_2020": ("Land Area (in square miles)", "2020 Census"),
        "population_2000": ("Population", "2000"),
        "population_current": ("Population", CURRENT_LABEL),
        "households_2000": ("Households", "2000"),
        "households_current": ("Households", CURRENT_LABEL),
        "housing_2000": ("Housing Units", "2000"),
        "housing_current": ("Housing Units", CURRENT_LABEL),
        "owner_current": ("Owner Occupied Units", CURRENT_LABEL),
        "renter_current": ("Renter-Occupied Units", CURRENT_LABEL),
        "vacant_current": ("Vacant Units", CURRENT_LABEL),
        "homeownership_current": ("Homeownership Rate", CURRENT_LABEL),
        "rent_2000": ("Median Gross Rent", "2000"),
        "rent_current": ("Median Gross Rent", CURRENT_LABEL),
        "home_value_2000": ("Median Home Value", "2000"),
        "home_value_current": ("Median Home Value", CURRENT_LABEL),
        "housing_cost_burden_current": ("% of Households Housing Cost-Burdened", CURRENT_LABEL),
        "household_income": ("Median Household Income", CURRENT_LABEL),
        "jobs_2023": ("Total Jobs in Municipality", "2023"),
    }
    missing = {key: pair for key, pair in required.items() if pair not in columns}
    if missing:
        raise RuntimeError(f"Neighborhood Trends source contract changed; missing columns: {missing}")
    c = {key: columns[pair] for key, pair in required.items()}

    municipalities, by_pair, unique = load_crosswalk()
    out: dict[str, dict] = {}
    unmatched: list[str] = []
    duplicate: list[str] = []

    for row in rows_iter:
        muni = clean_header(row[0] if len(row) > 0 else None)
        county = clean_header(row[1] if len(row) > 1 else None)
        if not muni:
            continue
        district = None
        for a in aliases(muni):
            district = by_pair.get((a, county_norm(county)))
            if district:
                break
        if not district:
            for a in aliases(muni):
                district = unique.get(a)
                if district:
                    break
        if not district:
            unmatched.append(f"{muni} ({county})")
            continue
        if district in out:
            duplicate.append(district)
            continue

        land = num(row[c["land_area_2020"]])
        jobs = num(row[c["jobs_2023"]])
        owner = num(row[c["owner_current"]])
        renter = num(row[c["renter_current"]])
        occupied = (owner + renter) if owner is not None and renter is not None else None
        housing = num(row[c["housing_current"]])
        source_homeownership = percent_value(row[c["homeownership_current"]])
        calculated_owner_share = pct_share(owner, occupied)
        # Require the source's published Homeownership Rate and the transparent
        # owner/occupied calculation to agree within normal source rounding.
        if source_homeownership is not None and calculated_owner_share is not None and abs(source_homeownership - calculated_owner_share) > 0.11:
            raise RuntimeError(
                f"Homeownership contract mismatch for {district}: source={source_homeownership}, calculated={calculated_owner_share}"
            )

        base = municipalities.get(district, {})
        record = {
            "district": district,
            "name": base.get("name") or muni,
            "county": base.get("county") or county,
            "population_change": pct_change(row[c["population_2000"]], row[c["population_current"]]),
            "housing_unit_change": pct_change(row[c["housing_2000"]], row[c["housing_current"]]),
            "rental_cost_change": pct_change(row[c["rent_2000"]], row[c["rent_current"]]),
            "home_value_change": pct_change(row[c["home_value_2000"]], row[c["home_value_current"]]),
            "household_income": rounded(row[c["household_income"]], 0),
            "employment_density": round(jobs / land, 2) if jobs is not None and land and land > 0 else None,
            "neighborhood_trend_year": 2024,
            # Exact/deterministic municipality facts used to satisfy the compatible
            # Municipal Housing Profile catalog semantics. These do not use Power BI
            # scraping; they come from this same official DCA municipality workbook.
            "housing_stock": rounded(housing, 0),
            "owner_occupied_share": calculated_owner_share,
            "renter_occupied_share": pct_share(renter, occupied),
            "vacancy_rate": pct_share(row[c["vacant_current"]], housing),
            "median_gross_rent": rounded(row[c["rent_current"]], 0),
            "median_home_value": rounded(row[c["home_value_current"]], 0),
            "housing_cost_burden_share": percent_value(row[c["housing_cost_burden_current"]]),
            "household_growth": pct_change(row[c["households_2000"]], row[c["households_current"]]),
        }
        out[district] = record

    wb.close()
    temp.unlink(missing_ok=True)

    if duplicate:
        raise RuntimeError(f"Duplicate municipality matches: {sorted(set(duplicate))[:20]}")
    if len(unmatched) > 5:
        raise RuntimeError(f"Too many unmatched municipalities ({len(unmatched)}): {unmatched[:20]}")

    payload = {
        "schema_version": 2,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "NJ DCA 2026 Neighborhood Trends Database",
        "source_url": SOURCE,
        "source_bytes": len(response.content),
        "source_window": "2000 Census / 2020 Census / 2020-24 ACS 5-Year Estimates / 2023 jobs",
        "field_contract": {
            "population_change": "percent change: 2000 population to 2020-24 estimate",
            "housing_unit_change": "percent change: 2000 housing units to 2020-24 estimate",
            "rental_cost_change": "percent change: 2000 median gross rent to 2020-24 estimate; nominal dollars",
            "home_value_change": "percent change: 2000 median home value to 2020-24 estimate; nominal dollars",
            "household_income": "2020-24 median household income, dollars",
            "employment_density": "2023 total jobs divided by 2020 Census land area, jobs per square mile",
            "neighborhood_trend_year": "2024, endpoint of current 2020-24 ACS estimate window",
            "housing_stock": "2020-24 Housing Units, municipality total",
            "owner_occupied_share": "Owner Occupied Units / (Owner Occupied Units + Renter-Occupied Units) × 100; cross-checked to published Homeownership Rate",
            "renter_occupied_share": "Renter-Occupied Units / (Owner Occupied Units + Renter-Occupied Units) × 100",
            "vacancy_rate": "Vacant Units / Housing Units × 100",
            "median_gross_rent": "2020-24 Median Gross Rent, dollars",
            "median_home_value": "2020-24 Median Home Value, dollars",
            "housing_cost_burden_share": "published 2020-24 % of Households Housing Cost-Burdened, percentage points",
            "household_growth": "percent change: 2000 Households to 2020-24 Households",
        },
        "municipal_housing_profile_compatibility": {
            "source_contract": "official DCA Neighborhood Trends municipality values, not Power BI extraction",
            "exact_or_deterministic_fields": [
                "housing_stock",
                "owner_occupied_share",
                "renter_occupied_share",
                "vacancy_rate",
                "median_gross_rent",
                "median_home_value",
                "housing_cost_burden_share",
                "household_growth",
            ],
            "excluded_fields": {
                "eviction_rate": "DCA Municipal Housing Profile documentation identifies eviction filtering at ZIP geography; not converted to municipality scope.",
                "housing_production": "Catalog label is broader than a single proven workbook field; no synthetic alias is emitted."
            }
        },
        "excluded_catalog_fields": {
            "commute_mode_mix": "No exact municipality-level commute-mode mix field in the 2026 workbook contract.",
            "real_estate_tax_median": "No exact municipality-level median real-estate-tax field in the 2026 workbook contract.",
            "walkability_score": "Workbook publishes percent walking to work, not the catalog's walkability-score semantic.",
        },
        "municipalities_matched": len(out),
        "unmatched": unmatched,
        "municipalities": out,
    }
    args.output.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(out)} municipalities to {args.output}; unmatched={len(unmatched)}")


if __name__ == "__main__":
    main()
