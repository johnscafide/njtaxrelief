#!/usr/bin/env python3
"""Extract NJ DCA Fourth Round municipal obligations from published Appendix A.

The methodology PDF is DCA's published final municipal table. This extractor uses
Appendix A pages only, reconciles every row against the official calculation workbook
for FIPS/DCA Municode identity, and refuses output unless all 564 municipalities match
uniquely. Workbook formulas remain lineage; published appendix values are the final
value contract.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import pdfplumber
import requests
from openpyxl import load_workbook

PDF_URL = "https://www.nj.gov/dca/dlps/pdf/FourthRoundCalculation_Methodology.pdf"
XLSX_URL = "https://www.nj.gov/dca/dlps/pdf/FourthRoundCalculation_Workbook.xlsx"
SOURCE_PAGE = "https://www.nj.gov/dca/dlps/4th_Round_Numbers.shtml"
EXPECTED = 564


def clean(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "").replace("\n", " ")).strip()


def norm_name(v: object) -> str:
    s = clean(v).upper().replace(".", "")
    s = re.sub(r"\bTOWNSHIP\b", "TWP", s)
    s = re.sub(r"\bBOROUGH\b", "BORO", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_county(v: object) -> str:
    return re.sub(r"\s+COUNTY$", "", norm_name(v))


def num(v: object):
    s = clean(v).replace(",", "").replace("%", "")
    if not s or s in {"-", "—", "N/A"}:
        return None
    try:
        x = float(s)
        return int(x) if x.is_integer() else round(x, 6)
    except ValueError:
        return None


def pct(v: object):
    return num(v)


def workbook_index(path: Path):
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb["Final Summary"]
    leaves = [clean(v) for v in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
    def col(name): return leaves.index(name)
    ci = {
        "fips": col("County Subdivision FIPS Code"),
        "dca_municode": col("DCA Municode"),
        "municipality": col("Municipality"),
        "county": col("County"),
        "region": col("Region"),
    }
    by_pair = defaultdict(list)
    rows = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        muni, county = clean(row[ci["municipality"]]), clean(row[ci["county"]])
        if not muni or not county:
            continue
        fips = re.sub(r"\D", "", clean(row[ci["fips"]]))
        if len(fips) == 10: fips = fips[-5:]
        dca = re.sub(r"\D", "", clean(row[ci["dca_municode"]]))
        rec = {
            "county_subdivision_fips": fips,
            "dca_municode": dca.zfill(4) if dca else None,
            "municipality": muni,
            "county": county,
            "region": num(row[ci["region"]]),
        }
        key = (norm_name(muni), norm_county(county))
        by_pair[key].append(rec)
        rows[fips] = rec
    if len(rows) != EXPECTED:
        raise RuntimeError(f"Workbook identity gate failed: {len(rows)} municipalities")
    return by_pair


def table_rows(pdf_path: Path):
    out = []
    raw_samples = []
    with pdfplumber.open(pdf_path) as pdf:
        for pageno in range(19, min(42, len(pdf.pages))):
            page = pdf.pages[pageno]
            tables = page.extract_tables({
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "intersection_tolerance": 5,
                "snap_tolerance": 4,
                "join_tolerance": 4,
            })
            for table in tables:
                for row in table or []:
                    cells = [clean(x) for x in (row or [])]
                    if len(cells) < 12:
                        continue
                    if cells[0].lower().startswith("municipality") or "Present Need" in cells[0]:
                        continue
                    muni, county = cells[0], cells[1]
                    if not muni or not county or not num(cells[2]):
                        continue
                    if len(raw_samples) < 12:
                        raw_samples.append({"page":pageno+1,"cell_count":len(cells),"cells":cells})
                    rec = {
                        "municipality": muni,
                        "county": county,
                        "region": num(cells[2]),
                        "present_need": num(cells[3]),
                        "qualified_urban_aid": cells[4] or None,
                        "nonresidential_value_factor_pct": pct(cells[5]),
                        "land_capacity_factor_pct": pct(cells[6]),
                        "income_capacity_factor_pct": pct(cells[7]),
                        "average_allocation_factor_pct": pct(cells[8]),
                        "prospective_need": num(cells[9]),
                        "cap_1000_20pct": num(cells[10]),
                        "prospective_need_capped": num(cells[11]),
                        "appendix_pdf_page": pageno + 1,
                    }
                    out.append(rec)
    return out, raw_samples


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", type=Path, required=True)
    args = ap.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    pdf_path = args.output.with_suffix(".pdf")
    xlsx_path = args.output.with_suffix(".xlsx")
    rp=requests.get(PDF_URL,timeout=180);rp.raise_for_status();pdf_path.write_bytes(rp.content)
    rx=requests.get(XLSX_URL,timeout=180);rx.raise_for_status();xlsx_path.write_bytes(rx.content)

    by_pair = workbook_index(xlsx_path)
    extracted, raw_samples = table_rows(pdf_path)
    matched = {}
    unmatched = []
    ambiguous = []
    duplicate = []
    for rec in extracted:
        key = (norm_name(rec["municipality"]), norm_county(rec["county"]))
        candidates = by_pair.get(key, [])
        if len(candidates) != 1:
            (ambiguous if len(candidates) > 1 else unmatched).append({"record": rec, "candidate_count": len(candidates)})
            continue
        ident = candidates[0]
        fips = ident["county_subdivision_fips"]
        if fips in matched:
            duplicate.append(fips)
            continue
        if rec["present_need"] is None or rec["prospective_need"] is None or rec["prospective_need_capped"] is None:
            unmatched.append({"record": rec, "reason": "missing published required value"})
            continue
        matched[fips] = {**ident, **rec}

    errors = []
    if len(extracted) != EXPECTED: errors.append(f"appendix_rows={len(extracted)} expected={EXPECTED}")
    if len(matched) != EXPECTED: errors.append(f"matched={len(matched)} expected={EXPECTED}")
    if unmatched: errors.append(f"unmatched={len(unmatched)}")
    if ambiguous: errors.append(f"ambiguous={len(ambiguous)}")
    if duplicate: errors.append(f"duplicates={len(duplicate)}")
    present_total = sum(x["present_need"] for x in matched.values()) if matched else 0
    if present_total != 65410: errors.append(f"present_total={present_total} expected=65410")

    payload = {
        "schema_version": 2,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "NJ DCA Fourth Round (2025-2035) Methodology Appendix A + Calculation Workbook identity crosswalk",
        "source_page": SOURCE_PAGE,"pdf_url":PDF_URL,"workbook_url":XLSX_URL,
        "legal_context": "DCA publishes these calculations as non-binding guidance. Watchdog presents observed DCA calculation values, not a legal determination of municipal obligation.",
        "validation": {
            "publishable": not errors,"errors":errors,"appendix_rows":len(extracted),"matched_municipalities":len(matched),
            "present_need_statewide_total":present_total,"raw_cell_samples":raw_samples,
            "unmatched_examples":unmatched[:20],"ambiguous_examples":ambiguous[:20],"duplicate_examples":duplicate[:20]
        },
        "municipalities": matched if not errors else {},
    }
    args.output.write_text(json.dumps(payload,separators=(",", ":")),encoding="utf-8")
    pdf_path.unlink(missing_ok=True);xlsx_path.unlink(missing_ok=True)
    print(json.dumps(payload["validation"],sort_keys=True))
    if errors:sys.exit(2)


if __name__ == "__main__":
    main()
